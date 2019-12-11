from openpyxl import *
wb = load_workbook('./NIR_Data_Hist.xls')

sheet_obj = wb.active
m_row = sheet_obj.max_row

# Выводим значения в цикле
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2)
print(cell_obj.value)